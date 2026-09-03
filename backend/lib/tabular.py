"""Read an uploaded leads table (.xlsx / .xls / .csv) into a header row plus string rows.

Every cell comes back as a trimmed string so the import path only ever reasons about one
shape: spreadsheet numbers lose their float tail (35.0 -> "35"), dates become YYYY-MM-DD,
and empty cells become "".
"""

import csv
import io
from datetime import date, datetime
from typing import List, Tuple

MAX_UPLOAD_BYTES = 5 * 1024 * 1024
SUPPORTED_SUFFIXES = (".csv", ".xlsx", ".xlsm", ".xls")


class TableError(ValueError):
    """Carries a Bahasa Indonesia message the API hands straight back to the admin."""


def _cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Ya" if value else "Tidak"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        # Spreadsheets hold every number as a float; a phone column must not become "8.1e+10".
        return str(int(value)) if value.is_integer() else repr(value)
    return str(value).strip()


def _rows_from_csv(raw: bytes) -> List[List[str]]:
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")
    try:
        # Indonesian Excel often exports semicolon-separated CSV.
        dialect: object = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return [[_cell(c) for c in row] for row in csv.reader(io.StringIO(text), dialect)]  # type: ignore[arg-type]


def _rows_from_xlsx(raw: bytes) -> List[List[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        sheet = wb[wb.sheetnames[0]]
        return [[_cell(c) for c in row] for row in sheet.iter_rows(values_only=True)]
    finally:
        wb.close()


def _rows_from_xls(raw: bytes) -> List[List[str]]:
    import xlrd

    book = xlrd.open_workbook(file_contents=raw)
    sheet = book.sheet_by_index(0)
    out: List[List[str]] = []
    for index in range(sheet.nrows):
        row: List[str] = []
        for cell in sheet.row(index):
            if cell.ctype == xlrd.XL_CELL_DATE:
                row.append(_cell(datetime(*xlrd.xldate_as_tuple(cell.value, book.datemode))))
            else:
                row.append(_cell(cell.value))
        out.append(row)
    return out


def _pick_reader(filename: str, raw: bytes):
    lower = (filename or "").lower()
    if lower.endswith((".xlsx", ".xlsm")):
        return _rows_from_xlsx
    if lower.endswith(".xls"):
        # Some exporters name a real .xlsx "report.xls"; trust the bytes over the extension.
        return _rows_from_xlsx if raw[:2] == b"PK" else _rows_from_xls
    if lower.endswith(".csv"):
        return _rows_from_csv
    if raw[:2] == b"PK":
        return _rows_from_xlsx
    if raw[:4] == b"\xd0\xcf\x11\xe0":
        return _rows_from_xls
    return _rows_from_csv


def parse_table(filename: str, raw: bytes) -> Tuple[List[str], List[List[str]]]:
    """Return (headers, data rows). Rows are padded to the header width."""
    if not raw:
        raise TableError("File kosong")
    if len(raw) > MAX_UPLOAD_BYTES:
        raise TableError("Ukuran file maksimal 5 MB")

    try:
        rows = _pick_reader(filename, raw)(raw)
    except TableError:
        raise
    except Exception as exc:  # noqa: BLE001 - any parser failure is the same user-facing story
        raise TableError(
            f"File tidak bisa dibaca ({type(exc).__name__}). Gunakan format .xlsx, .xls, atau .csv."
        ) from exc

    rows = [r for r in rows if any(c for c in r)]
    if not rows:
        raise TableError("File tidak berisi data")

    raw_headers = rows[0]
    headers: List[str] = []
    for position, name in enumerate(raw_headers, start=1):
        label = name or f"Kolom {position}"
        # Duplicate headers would silently overwrite each other once zipped into a dict.
        while label in headers:
            label = f"{label} ({position})"
        headers.append(label)

    body = [(r + [""] * len(headers))[: len(headers)] for r in rows[1:]]
    if not body:
        raise TableError("File hanya berisi baris header, belum ada data")
    return headers, body
