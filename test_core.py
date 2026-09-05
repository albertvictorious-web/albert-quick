"""POC fondasi deploy Vercel + MongoDB Atlas.

Membuktikan 4 hal sebelum app dipindah:
  1. Koneksi ke MongoDB Atlas berhasil (TLS/SRV resolve dari environment ini)
  2. Write + read + delete berhasil (user database punya hak readWrite)
  3. Pola "cached client" (yang dipakai serverless Vercel) tidak bocor koneksi
  4. bcrypt hash + JWT encode/decode jalan (dependency auth yang akan dibundel Vercel)

Kredensial TIDAK ditulis di file ini. Diambil dari backend/.env (MONGO_URL), atau
override lewat environment: ATLAS_TEST_URI="mongodb+srv://..." python test_core.py
"""

import asyncio
import os
import sys
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path

from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / "backend" / ".env")

ATLAS_URI = os.environ.get("ATLAS_TEST_URI") or os.environ.get("MONGO_URL", "")
DB_NAME = os.environ.get("ATLAS_TEST_DB") or "quickpro_crm"

if not ATLAS_URI:
    sys.exit(
        "MONGO_URL tidak ditemukan. Isi backend/.env (lihat .env.example) atau "
        'jalankan: ATLAS_TEST_URI="mongodb+srv://..." python test_core.py'
    )

results = []


def record(name, ok, detail=""):
    results.append((name, ok, detail))
    print(f"[{'PASS' if ok else 'FAIL'}] {name}" + (f" — {detail}" if detail else ""))


# --- cached client pattern: persis yang akan dipakai di serverless -----------------
_cached = {}


def get_db():
    """Satu client per proses (warm container Vercel reuse ini, bukan bikin baru)."""
    from motor.motor_asyncio import AsyncIOMotorClient

    if "client" not in _cached:
        _cached["client"] = AsyncIOMotorClient(
            ATLAS_URI,
            maxPoolSize=5,
            serverSelectionTimeoutMS=15000,
            tz_aware=True,
        )
    return _cached["client"][DB_NAME]


async def test_1_connect():
    try:
        db = get_db()
        info = await db.command("ping")
        record("1. Ping Atlas", info.get("ok") == 1.0, f"ok={info.get('ok')}")
        build = await db.client.admin.command("buildInfo")
        record("1b. Versi server MongoDB", True, f"MongoDB {build['version']}")
        return True
    except Exception as e:
        record("1. Ping Atlas", False, f"{type(e).__name__}: {e}")
        return False


async def test_2_crud():
    try:
        db = get_db()
        marker = f"poc-{uuid.uuid4().hex[:8]}"
        doc = {
            "id": marker,
            "nama": "POC Lead",
            "created_at": datetime.now(timezone.utc),
            "nested": {"angka": 42, "list": [1, 2, 3]},
        }
        await db.poc_check.insert_one(doc)
        found = await db.poc_check.find_one({"id": marker})
        record("2a. Insert + find", found is not None and found["nama"] == "POC Lead")

        await db.poc_check.update_one({"id": marker}, {"$set": {"nama": "POC Updated"}})
        found = await db.poc_check.find_one({"id": marker})
        record("2b. Update", found["nama"] == "POC Updated")

        # datetime harus kembali sebagai datetime (tz_aware) — dipakai rekap bulanan
        record("2c. Datetime round-trip tz-aware", found["created_at"].tzinfo is not None,
               str(found["created_at"]))

        res = await db.poc_check.delete_many({"id": marker})
        record("2d. Delete (bersihkan)", res.deleted_count == 1)
        return True
    except Exception as e:
        record("2. CRUD Atlas", False, f"{type(e).__name__}: {e}")
        return False


async def test_3_index_and_reuse():
    try:
        db = get_db()
        # create_index butuh hak yang sama dengan yang app butuh saat startup
        name = await db.poc_check.create_index("id")
        record("3a. Create index (hak readWrite)", bool(name), f"index={name}")
        await db.poc_check.drop_index(name)

        # client di-reuse, bukan dibuat ulang -> syarat serverless
        db2 = get_db()
        record("3b. Client di-reuse antar panggilan", db2.client is db.client)
        return True
    except Exception as e:
        record("3. Index / reuse", False, f"{type(e).__name__}: {e}")
        return False


async def test_4_concurrent():
    """Serverless bisa hit banyak request paralel di satu container."""
    try:
        db = get_db()
        async def one(i):
            await db.poc_check.insert_one({"id": f"conc-{i}", "i": i})
            return await db.poc_check.find_one({"id": f"conc-{i}"})

        docs = await asyncio.gather(*[one(i) for i in range(10)])
        ok = all(d is not None for d in docs)
        await db.poc_check.delete_many({"id": {"$regex": "^conc-"}})
        record("4. 10 operasi paralel di 1 pool", ok)
        return ok
    except Exception as e:
        record("4. Operasi paralel", False, f"{type(e).__name__}: {e}")
        return False


def test_5_auth_deps():
    """bcrypt + PyJWT: dua dependency native yang paling sering gagal di Vercel."""
    try:
        from passlib.context import CryptContext

        ctx = CryptContext(schemes=["bcrypt"], deprecated="auto")
        h = ctx.hash("admin123")
        record("5a. bcrypt hash + verify", ctx.verify("admin123", h) and not ctx.verify("salah", h))
    except Exception as e:
        record("5a. bcrypt", False, f"{type(e).__name__}: {e}")

    try:
        import jwt

        tok = jwt.encode(
            {"sub": "abc", "exp": datetime.now(timezone.utc) + timedelta(days=7)},
            "secret-poc",
            algorithm="HS256",
        )
        payload = jwt.decode(tok, "secret-poc", algorithms=["HS256"])
        record("5b. JWT encode + decode", payload["sub"] == "abc")
    except Exception as e:
        record("5b. JWT", False, f"{type(e).__name__}: {e}")


def test_6_excel_deps():
    """openpyxl/xlrd dipakai fitur Upload Leads — pastikan ikut jalan tanpa pandas."""
    try:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nama Lengkap", "No. HP / WhatsApp", "Umur"])
        ws.append(["Budi POC", "081200000001", 30])
        import io

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        wb2 = openpyxl.load_workbook(buf)
        rows = list(wb2.active.iter_rows(values_only=True))
        record("6. openpyxl baca/tulis xlsx (tanpa pandas)", rows[1][0] == "Budi POC", str(rows[1]))
    except Exception as e:
        record("6. openpyxl", False, f"{type(e).__name__}: {e}")


async def main():
    print("=" * 70)
    print("POC FONDASI: MongoDB Atlas + dependency serverless Vercel")
    print("=" * 70)
    ok = await test_1_connect()
    if ok:
        await test_2_crud()
        await test_3_index_and_reuse()
        await test_4_concurrent()
    test_5_auth_deps()
    test_6_excel_deps()

    if "client" in _cached:
        _cached["client"].close()

    print("=" * 70)
    passed = sum(1 for _, o, _ in results if o)
    total = len(results)
    print(f"HASIL: {passed}/{total} lulus")
    failed = [n for n, o, _ in results if not o]
    if failed:
        print("GAGAL:", ", ".join(failed))
        sys.exit(1)
    print("SEMUA LULUS — fondasi Atlas + dependency siap untuk Vercel")


if __name__ == "__main__":
    asyncio.run(main())
